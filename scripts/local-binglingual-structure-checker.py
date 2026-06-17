#!/usr/bin/env python3
"""
Check local bilingual Markdown document structures.

Edit the configuration variables below, then run:

    python local-binglingual-structure-checker.py
"""

from __future__ import annotations

from datetime import datetime
import json
from pathlib import Path
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from resolve_cloud_source_files import extract_markdown_doc_links
from translation_structure_validator import (
    HEADING_RE,
    StructureValidationIssue,
    extract_custom_content_tags,
    iter_markdown_content_lines,
    validate_markdown_heading_structures,
)


# ---------------------------------------------------------------------------
# Configuration — edit these variables before running
# ---------------------------------------------------------------------------

SOURCE_LANGUAGE = "English"
TARGET_LANGUAGE = "Japanese" ## Options: "Chinese", "Japanese"

SOURCE_ROOT = "/Users/grcai/Documents/GitHub/docs-en-release-8.5"
TARGET_ROOT = "/Users/grcai/Documents/GitHub/docs"

TOC_SCOPE = "all"  # Options: "all", "partial". "all" means all TOC*.md files in SOURCE_ROOT will be scanned, while "partial" means only the files listed in TOC_LIST will be scanned.

TOC_LIST = [
    "TOC-tidb-cloud.md",
    "TOC-tidb-cloud-starter.md",
    "TOC-tidb-cloud-essential.md",
    "TOC-tidb-cloud-premium.md",
    "TOC-tidb-cloud-releases.md",
]

# Maximum issues to print in the terminal; 0 means print all.
PRINT_LIMIT = 0

# Optional output paths; leave empty to skip JSON or use a default Excel path.
JSON_OUT = ""
EXCEL_OUT = ""  # defaults to local_structure_check_<timestamp>.xlsx next to this script


# ---------------------------------------------------------------------------
# Implementation — no need to edit below
# ---------------------------------------------------------------------------

def normalize_doc_path(value: str) -> str:
    rel = (value or "").strip().replace("\\", "/")
    rel = rel.split("#", 1)[0].split("?", 1)[0].strip()
    rel = rel.lstrip("/")
    if rel.startswith("./"):
        rel = rel[2:]
    if rel.startswith("docs/"):
        rel = rel[5:]
    return rel


def collect_toc_markdown_files(source_root: Path, toc_files: list[str]) -> list[str]:
    files = set()

    for toc_file in toc_files:
        toc_path = source_root / toc_file
        if not toc_path.exists():
            raise FileNotFoundError(f"TOC file not found: {toc_path}")

        for link in extract_markdown_doc_links(toc_path.read_text(encoding="utf-8")):
            rel = normalize_doc_path(link)
            if rel.endswith(".md"):
                files.add(rel)

    return sorted(files)


def discover_all_tocs(source_root: Path) -> list[str]:
    """Return sorted list of TOC*.md filenames found in source_root."""
    return sorted(p.name for p in source_root.glob("TOC*.md"))


def read_text(path: Path) -> str | None:
    if not path.exists():
        return None
    return path.read_text(encoding="utf-8")


def check_file(source_root: Path, target_root: Path, rel_path: str) -> list[StructureValidationIssue]:
    source_content = read_text(source_root / rel_path)
    target_content = read_text(target_root / rel_path)

    if source_content is None:
        return [StructureValidationIssue(rel_path, "source file missing")]
    if target_content is None:
        return [StructureValidationIssue(rel_path, "target file missing")]

    return validate_markdown_heading_structures(
        [rel_path],
        source_content_loader=lambda _: source_content,
        target_content_loader=lambda _: target_content,
    )


def issue_category(issue: StructureValidationIssue) -> str:
    reason = issue.reason.lower()
    if "customcontent" in reason:
        return "CustomContent"
    if "heading" in reason:
        return "Heading"
    if "missing" in reason:
        return "Missing file"
    return "Structure"


_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_ISSUE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_MISSING_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_FILE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_FILE_FONT = Font(bold=True, size=11, color="1F4E79")
_MATCH_FONT = Font(bold=True, color="006100")
_MISMATCH_FONT = Font(bold=True, color="9C0006")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_ISSUE_TYPE_ORDER = ["Heading", "CustomContent", "Missing file", "Structure"]


def default_excel_path() -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path(__file__).resolve().parent / f"local_structure_check_{timestamp}.xlsx"


def _group_issues_by_type(issues: list[StructureValidationIssue]):
    groups = {}
    for issue in issues:
        groups.setdefault(issue_category(issue), []).append(issue)

    ordered = [
        (issue_type, groups.pop(issue_type))
        for issue_type in _ISSUE_TYPE_ORDER
        if issue_type in groups
    ]
    ordered.extend(sorted(groups.items()))
    return ordered


def _extract_heading_details(content: str | None):
    if content is None:
        return []

    headings = []
    for line_number, line in iter_markdown_content_lines(content):
        match = HEADING_RE.match(line)
        if match:
            headings.append({
                "line": line_number,
                "level": len(match.group(1)),
                "text": match.group(2).strip(),
            })
    return headings


def _format_heading_details(content: str | None) -> str:
    if content is None:
        return "(file missing or unavailable)"

    headings = _extract_heading_details(content)
    if not headings:
        return "(no headings)"
    return "\n".join(
        f"line {item['line']}: {'#' * item['level']} {item['text']}"
        for item in headings
    )


def _format_custom_content_details(content: str | None) -> str:
    if content is None:
        return "(file missing or unavailable)"
    tags = extract_custom_content_tags(content)
    if not tags:
        return "0 CustomContent tags"
    return "\n".join(f"line {tag.line_number}: {tag.text}" for tag in tags)


def _read_issue_content(root: Path | None, rel_path: str, cache: dict[tuple[Path, str], str | None]):
    if root is None:
        return None
    key = (root, rel_path)
    if key not in cache:
        cache[key] = read_text(root / rel_path)
    return cache[key]


def _style_header_row(worksheet, row_num: int, headers: list[str]) -> None:
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=row_num, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    worksheet.row_dimensions[row_num].height = 32


def _write_file_summary_row(
    worksheet,
    row_num: int,
    issue: StructureValidationIssue,
    max_col: int,
) -> None:
    values = [
        issue.file_path,
        issue.reason,
        issue.first_difference,
        issue.source_compact,
        issue.target_compact,
    ]
    for col in range(1, max_col + 1):
        value = values[col - 1] if col <= len(values) else ""
        cell = worksheet.cell(row=row_num, column=col, value=value)
        cell.font = _FILE_FONT if col == 1 else Font(italic=True, size=10, color="404040")
        cell.fill = _FILE_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    worksheet.row_dimensions[row_num].height = 42


def _cell_value(item, key: str):
    if not item:
        return ""
    return item[key]


def _heading_level(item) -> str:
    if not item:
        return "(missing)"
    return "#" * item["level"]


def _heading_text(item) -> str:
    if not item:
        return "(missing)"
    return item["text"]


def _tag_line(tag):
    return tag.line_number if tag else ""


def _tag_text(tag) -> str:
    return tag.text if tag else "(missing)"


def _tag_signature(tag):
    return (tag.kind, tag.text) if tag else None


def _write_heading_issue_table(
    worksheet, row_num: int, issue, source_content, target_content,
    source_label: str = "English", target_label: str = "Chinese",
) -> int:
    headers = [
        "#",
        f"{source_label} Line",
        f"{source_label} Level",
        f"{source_label} Heading",
        f"{target_label} Line",
        f"{target_label} Level",
        f"{target_label} Heading",
        "Level Match",
    ]
    _write_file_summary_row(worksheet, row_num, issue, len(headers))
    row_num += 1
    _style_header_row(worksheet, row_num, headers)
    row_num += 1

    source_headings = _extract_heading_details(source_content)
    target_headings = _extract_heading_details(target_content)
    max_len = max(len(source_headings), len(target_headings))
    for index in range(max_len):
        src_heading = source_headings[index] if index < len(source_headings) else None
        tgt_heading = target_headings[index] if index < len(target_headings) else None
        levels_match = (
            src_heading is not None
            and tgt_heading is not None
            and src_heading["level"] == tgt_heading["level"]
        )
        values = [
            index + 1,
            _cell_value(src_heading, "line"),
            _heading_level(src_heading),
            _heading_text(src_heading),
            _cell_value(tgt_heading, "line"),
            _heading_level(tgt_heading),
            _heading_text(tgt_heading),
        ]
        fill = _MATCH_FILL if levels_match else _ISSUE_FILL
        if src_heading is None or tgt_heading is None:
            fill = _MISSING_FILL
        for col, value in enumerate(values, 1):
            cell = worksheet.cell(row=row_num, column=col, value=value)
            cell.fill = fill
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col in (1, 2, 3, 5, 6) else "left",
                vertical="top",
                wrap_text=True,
            )
        match_cell = worksheet.cell(row=row_num, column=len(headers))
        match_cell.value = "Match" if levels_match else "Mismatch"
        match_cell.font = _MATCH_FONT if levels_match else _MISMATCH_FONT
        match_cell.fill = fill
        match_cell.border = _THIN_BORDER
        match_cell.alignment = Alignment(horizontal="center", vertical="top")
        row_num += 1

    if max_len == 0:
        worksheet.cell(row=row_num, column=1, value="No headings found in either file.")
        row_num += 1

    return row_num + 1


def _write_custom_content_issue_table(
    worksheet, row_num: int, issue, source_content, target_content,
    source_label: str = "English", target_label: str = "Chinese",
) -> int:
    headers = [
        "#",
        f"{source_label} Line",
        f"{source_label} Tag",
        f"{target_label} Line",
        f"{target_label} Tag",
        "Tag Match",
    ]
    _write_file_summary_row(worksheet, row_num, issue, len(headers))
    row_num += 1
    _style_header_row(worksheet, row_num, headers)
    row_num += 1

    source_tags = extract_custom_content_tags(source_content)
    target_tags = extract_custom_content_tags(target_content)
    max_len = max(len(source_tags), len(target_tags))
    for index in range(max_len):
        src_tag = source_tags[index] if index < len(source_tags) else None
        tgt_tag = target_tags[index] if index < len(target_tags) else None
        tags_match = _tag_signature(src_tag) == _tag_signature(tgt_tag)
        values = [
            index + 1,
            _tag_line(src_tag),
            _tag_text(src_tag),
            _tag_line(tgt_tag),
            _tag_text(tgt_tag),
        ]
        fill = _MATCH_FILL if tags_match else _ISSUE_FILL
        if src_tag is None or tgt_tag is None:
            fill = _MISSING_FILL
        for col, value in enumerate(values, 1):
            cell = worksheet.cell(row=row_num, column=col, value=value)
            cell.fill = fill
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col in (1, 2, 4) else "left",
                vertical="top",
                wrap_text=True,
            )
        match_cell = worksheet.cell(row=row_num, column=len(headers))
        match_cell.value = "Match" if tags_match else "Mismatch"
        match_cell.font = _MATCH_FONT if tags_match else _MISMATCH_FONT
        match_cell.fill = fill
        match_cell.border = _THIN_BORDER
        match_cell.alignment = Alignment(horizontal="center", vertical="top")
        row_num += 1

    if max_len == 0:
        worksheet.cell(row=row_num, column=1, value="0 CustomContent tags in both files.")
        row_num += 1

    return row_num + 1


def _write_generic_issue_table(
    worksheet, row_num: int, issue, source_content, target_content,
    source_label: str = "English", target_label: str = "Chinese",
) -> int:
    headers = ["Field", source_label, target_label]
    _write_file_summary_row(worksheet, row_num, issue, len(headers))
    row_num += 1
    _style_header_row(worksheet, row_num, headers)
    row_num += 1

    rows = [
        ("Reason", issue.reason, issue.reason),
        ("First Difference", issue.first_difference, issue.first_difference),
        ("Structure", issue.source_compact, issue.target_compact),
        ("Headings", _format_heading_details(source_content), _format_heading_details(target_content)),
        (
            "CustomContent",
            _format_custom_content_details(source_content),
            _format_custom_content_details(target_content),
        ),
    ]
    for field, src_value, tgt_value in rows:
        for col, value in enumerate([field, src_value, tgt_value], 1):
            cell = worksheet.cell(row=row_num, column=col, value=value)
            cell.fill = _ISSUE_FILL
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        max_lines = max(str(value).count("\n") + 1 for value in (field, src_value, tgt_value) if value)
        worksheet.row_dimensions[row_num].height = min(220, max(36, max_lines * 15))
        row_num += 1

    return row_num + 1


def _write_issue_sheet(
    worksheet,
    issue_type: str,
    issues: list[StructureValidationIssue],
    source_root: Path | None,
    target_root: Path | None,
    content_cache: dict[tuple[Path, str], str | None],
    source_label: str = "English",
    target_label: str = "Chinese",
) -> None:
    row_num = 1
    for issue in issues:
        source_content = _read_issue_content(source_root, issue.file_path, content_cache)
        target_content = _read_issue_content(target_root, issue.file_path, content_cache)
        if issue_type == "Heading":
            row_num = _write_heading_issue_table(
                worksheet, row_num, issue, source_content, target_content,
                source_label, target_label,
            )
        elif issue_type == "CustomContent":
            row_num = _write_custom_content_issue_table(
                worksheet, row_num, issue, source_content, target_content,
                source_label, target_label,
            )
        else:
            row_num = _write_generic_issue_table(
                worksheet, row_num, issue, source_content, target_content,
                source_label, target_label,
            )

    widths_by_type = {
        "Heading": [5, 12, 14, 64, 12, 14, 64, 14],
        "CustomContent": [5, 12, 72, 12, 72, 14],
    }
    widths = widths_by_type.get(issue_type, [24, 72, 72])
    for col, width in enumerate(widths, 1):
        worksheet.column_dimensions[get_column_letter(col)].width = width


def write_excel_report(
    issues: list[StructureValidationIssue],
    output_path: Path,
    source_root: Path | None = None,
    target_root: Path | None = None,
    source_label: str = "English",
    target_label: str = "Chinese",
) -> None:
    """Write a local structure report that contains only problematic rows."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    grouped_issues = _group_issues_by_type(issues)
    if not grouped_issues:
        worksheet = workbook.active
        worksheet.title = "No Issues"
        worksheet.cell(row=1, column=1, value="No structure issues found.")
        workbook.save(output_path)
        return

    content_cache = {}
    for index, (issue_type, issue_list) in enumerate(grouped_issues):
        worksheet = workbook.active if index == 0 else workbook.create_sheet()
        worksheet.title = issue_type
        _write_issue_sheet(
            worksheet, issue_type, issue_list,
            source_root, target_root, content_cache,
            source_label, target_label,
        )

    workbook.save(output_path)


def main() -> int:
    source_root = Path(SOURCE_ROOT).expanduser()
    target_root = Path(TARGET_ROOT).expanduser()
    source_label = SOURCE_LANGUAGE
    target_label = TARGET_LANGUAGE

    if TOC_SCOPE == "all":
        toc_files = discover_all_tocs(source_root)
        if not toc_files:
            print(f"ERROR: no TOC*.md files found in {source_root}", file=sys.stderr)
            return 2
    else:
        toc_files = list(TOC_LIST)

    try:
        file_paths = collect_toc_markdown_files(source_root, toc_files)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2

    issues = []
    for rel_path in file_paths:
        issues.extend(check_file(source_root, target_root, rel_path))

    print(f"Source root ({source_label}): {source_root}")
    print(f"Target root ({target_label}): {target_root}")
    print(f"TOC scope: {TOC_SCOPE} ({len(toc_files)} TOC files: {', '.join(toc_files)})")
    print(f"Markdown files in scope: {len(file_paths)}")
    print(f"Document structure issues: {len(issues)}")

    if issues:
        print()
        print("Structure issues:")
        printable = issues if PRINT_LIMIT <= 0 else issues[:PRINT_LIMIT]
        for issue in printable:
            print(f"- {issue.file_path}: {issue.reason}")
            if issue.first_difference:
                print(f"  first difference: {issue.first_difference}")
            if issue.source_compact:
                print(f"  {source_label}: {issue.source_compact}")
            if issue.target_compact:
                print(f"  {target_label}: {issue.target_compact}")

        if PRINT_LIMIT > 0 and len(issues) > PRINT_LIMIT:
            print(f"... {len(issues) - PRINT_LIMIT} more issue(s) omitted by PRINT_LIMIT")

    if JSON_OUT:
        out_path = Path(JSON_OUT).expanduser()
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(
            json.dumps([issue.to_dict() for issue in issues], ensure_ascii=False, indent=2)
            + "\n",
            encoding="utf-8",
        )
        print()
        print(f"JSON report written to: {out_path}")

    excel_path = Path(EXCEL_OUT).expanduser() if EXCEL_OUT else default_excel_path()
    write_excel_report(issues, excel_path, source_root, target_root, source_label, target_label)
    print()
    print(f"Excel report written to: {excel_path}")

    return 1 if issues else 0


if __name__ == "__main__":
    raise SystemExit(main())
