#!/usr/bin/env python3
"""
Verify translation quality by comparing source diff and target PR file changes.

Fetches file-level change statistics (additions / deletions) from a source
diff (commit compare or PR) and a target PR, then outputs an Excel report
highlighting markdown files whose line-change counts diverge beyond a
configurable threshold.

Usage:
    export GITHUB_TOKEN="ghp_..."
    python verify_translation.py
"""

import os
import re
import subprocess
import sys
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

try:
    from github import Github, Auth
except ImportError:
    sys.exit("PyGithub is required: pip install PyGithub")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

from translation_structure_validator import (
    compare_added_file_line_integrity,
    compare_custom_content_structure,
    compact_custom_content_tags,
    extract_custom_content_tags,
    strip_related_resources_sections,
)


# ── Configuration ────────────────────────────────────────────────────────────
# Mode: "commit-based" uses a compare URL; "pr" uses a source PR URL.
MODE = "commit-based"

SOURCE_COMMIT_COMPARE = (
    "https://github.com/pingcap/docs/compare/2eaf0b7cd9c870d6f25c0dea7c7e1bb64ba2572b...6376b9086871957f4809711c83e021a01b0a8f5d"
)
SOURCE_PR = ""
TARGET_PR = "https://github.com/pingcap/docs/pull/23099"

# Optional: local clone of the source repo.  When set, `git diff --numstat`
# is used instead of the GitHub compare API, which avoids the 300-file cap.
SOURCE_REPO_PATH = "/Users/grcai/Documents/GitHub/docs"

WARNING_THRESHOLD = 5

# Files to exclude from the report entirely (expected artifacts, not translations).
EXCLUDE_FILES = {"latest_translation_commit.json"}

GITHUB_TOKEN = os.getenv("GITHUB_TOKEN", "")


def _positive_int_env(name, default):
    try:
        value = int(os.getenv(name, str(default)))
    except (TypeError, ValueError):
        return default
    return max(1, value)


DOCUMENT_STRUCTURE_WORKERS = _positive_int_env(
    "DOCUMENT_STRUCTURE_WORKERS",
    _positive_int_env("CUSTOM_CONTENT_WORKERS", 3),
)


# ── URL Parsers ──────────────────────────────────────────────────────────────

def parse_pr_url(pr_url):
    """Extract (owner, repo, pr_number) from a GitHub PR URL."""
    parts = pr_url.rstrip("/").split("/")
    return parts[-4], parts[-3], int(parts[-1])


def parse_compare_url(url):
    """Extract (owner, repo, base_ref, head_ref) from a GitHub compare URL."""
    parts = url.rstrip("/").split("/")
    owner, repo = parts[-4], parts[-3]
    base, head = parts[-1].split("...")
    return owner, repo, base, head


# ── Data Collection ──────────────────────────────────────────────────────────

def collect_file_stats(files):
    """Build a dict of {filename: stats} from GitHub file objects."""
    stats = {}
    for f in files:
        noop_pairs = _count_noop_line_change_pairs(getattr(f, "patch", "") or "")
        additions = max(0, f.additions - noop_pairs)
        deletions = max(0, f.deletions - noop_pairs)
        stats[f.filename] = {
            "status": f.status,
            "additions": additions,
            "deletions": deletions,
            "changes": additions + deletions,
            "is_md": f.filename.lower().endswith(".md"),
        }
    return stats


def _count_noop_line_change_pairs(patch):
    """Count changed line pairs that do not change content.

    Git can report an unchanged final line as one deletion plus one addition
    when only the EOF newline changed.  Those pairs are noise for translation
    line-count comparison, so subtract them from both sides.
    """
    noop_pairs = 0
    removed = []
    added = []

    def flush_change_block():
        nonlocal noop_pairs, removed, added
        if removed and added:
            removed_counts = Counter(removed)
            added_counts = Counter(added)
            noop_pairs += sum(
                min(count, added_counts[line])
                for line, count in removed_counts.items()
            )
        removed = []
        added = []

    for line in patch.splitlines():
        if line.startswith("@@"):
            flush_change_block()
            continue
        if line.startswith("\\"):
            continue
        if line.startswith("-") and not line.startswith("--- "):
            removed.append(line[1:])
            continue
        if line.startswith("+") and not line.startswith("+++ "):
            added.append(line[1:])
            continue
        flush_change_block()

    flush_change_block()
    return noop_pairs


def _normalize_diff_header_path(path):
    path = path.strip()
    if path == "/dev/null":
        return path
    if path.startswith("a/") or path.startswith("b/"):
        return path[2:]
    return path


def _split_git_diff_patches(diff_output):
    """Return {filename: patch_text} from a `git diff --patch` output."""
    patches = {}
    current_lines = []
    old_path = None
    new_path = None

    def flush_file():
        nonlocal current_lines, old_path, new_path
        if not current_lines:
            return
        filepath = new_path if new_path and new_path != "/dev/null" else old_path
        if filepath and filepath != "/dev/null":
            patches[filepath] = "\n".join(current_lines)
        current_lines = []
        old_path = None
        new_path = None

    for line in diff_output.splitlines():
        if line.startswith("diff --git "):
            flush_file()
        current_lines.append(line)
        if line.startswith("--- "):
            old_path = _normalize_diff_header_path(line[4:])
        elif line.startswith("+++ "):
            new_path = _normalize_diff_header_path(line[4:])

    flush_file()
    return patches


def _apply_noop_line_change_adjustments(stats, patches):
    for filename, patch in patches.items():
        if filename not in stats:
            continue
        noop_pairs = _count_noop_line_change_pairs(patch)
        if not noop_pairs:
            continue
        additions = max(0, stats[filename]["additions"] - noop_pairs)
        deletions = max(0, stats[filename]["deletions"] - noop_pairs)
        stats[filename]["additions"] = additions
        stats[filename]["deletions"] = deletions
        stats[filename]["changes"] = additions + deletions


def _git_numstat(repo_path, base, head):
    """Run `git diff --numstat` locally and return a stats dict.

    Each entry maps filename -> {status, additions, deletions, changes, is_md}.
    Binary files (shown as '-' in numstat) get additions/deletions = 0.
    """
    name_status_output = subprocess.check_output(
        ["git", "-C", repo_path, "diff", "--name-status", f"{base}...{head}"],
        text=True,
    )
    status_map = {}
    for line in name_status_output.strip().splitlines():
        parts = line.split("\t")
        if not parts or not parts[0]:
            continue
        code = parts[0][0]
        if code == "R" and len(parts) >= 3:
            status_map[parts[2]] = "renamed"
        elif code == "A":
            status_map[parts[1]] = "added"
        elif code == "D":
            status_map[parts[1]] = "removed"
        else:
            status_map[parts[-1]] = "modified"

    numstat_output = subprocess.check_output(
        ["git", "-C", repo_path, "diff", "--numstat", f"{base}...{head}"],
        text=True,
    )
    stats = {}
    for line in numstat_output.strip().splitlines():
        parts = line.split("\t")
        if len(parts) < 3:
            continue
        add_str, del_str, filename = parts[0], parts[1], parts[2]
        if "=>" in filename:
            filename = filename.split("=>")[-1].strip().rstrip("}")
            filename = filename.lstrip()
        additions = int(add_str) if add_str != "-" else 0
        deletions = int(del_str) if del_str != "-" else 0
        stats[filename] = {
            "status": status_map.get(filename, "modified"),
            "additions": additions,
            "deletions": deletions,
            "changes": additions + deletions,
            "is_md": filename.lower().endswith(".md"),
        }
    patch_output = subprocess.check_output(
        ["git", "-C", repo_path, "diff", "--unified=0", f"{base}...{head}"],
        text=True,
    )
    _apply_noop_line_change_adjustments(stats, _split_git_diff_patches(patch_output))
    return stats


def get_source_stats(github_client):
    """Fetch file change stats from the source diff (compare or PR)."""
    if MODE == "commit-based":
        owner, repo, base, head = parse_compare_url(SOURCE_COMMIT_COMPARE)
        label = f"{owner}/{repo} compare {base[:10]}...{head[:10]}"

        if SOURCE_REPO_PATH and os.path.isdir(SOURCE_REPO_PATH):
            print(f"  Source (local git diff): {label}")
            stats = _git_numstat(SOURCE_REPO_PATH, base, head)
            print(f"  Files in diff: {len(stats)}")
            return stats, label

        repository = github_client.get_repo(f"{owner}/{repo}")
        comparison = repository.compare(base, head)
        print(f"  Source (compare API): {label}")
        print(f"  Files in compare: {len(comparison.files)}")
        if len(comparison.files) >= 300:
            print("  ⚠  GitHub compare API caps at 300 files; results may be truncated.")
            print("     Set SOURCE_REPO_PATH to a local clone for accurate results.")
        return collect_file_stats(comparison.files), label
    else:
        owner, repo, pr_num = parse_pr_url(SOURCE_PR)
        repository = github_client.get_repo(f"{owner}/{repo}")
        pr = repository.get_pull(pr_num)
        files = list(pr.get_files())
        label = f"{owner}/{repo}#{pr_num} – {pr.title}"
        print(f"  Source (PR): {label}")
        print(f"  Files in PR: {len(files)}")
        return collect_file_stats(files), label


def get_target_stats(github_client):
    """Fetch file change stats from the target PR."""
    owner, repo, pr_num = parse_pr_url(TARGET_PR)
    repository = github_client.get_repo(f"{owner}/{repo}")
    pr = repository.get_pull(pr_num)
    files = list(pr.get_files())
    label = f"{owner}/{repo}#{pr_num} – {pr.title}"
    print(f"  Target (PR): {label}")
    print(f"  Files in PR: {len(files)}")
    return collect_file_stats(files), label


# ── Heading Structure Analysis ────────────────────────────────────────────────

def _extract_headings(content):
    """Extract markdown headings while skipping code blocks.

    Returns a list of ``(level, text)`` tuples in document order.
    """
    headings = []
    in_code_block = False
    for line in content.splitlines():
        stripped = line.strip()
        if stripped.startswith("```"):
            in_code_block = not in_code_block
            continue
        if in_code_block:
            continue
        if line != line.lstrip():
            continue
        m = re.match(r"^(#{1,6})\s+(.+)", line)
        if m:
            headings.append((len(m.group(1)), m.group(2).strip()))
    return headings


def _compact_structure(headings):
    """Run-length encode heading levels.

    Example: [(1,'A'),(2,'B'),(2,'C'),(3,'D')] -> ``'#×1 ##×2 ###×1'``
    """
    if not headings:
        return "(no headings)"
    levels = [h[0] for h in headings]
    runs, i = [], 0
    while i < len(levels):
        lv, count = levels[i], 1
        while i + count < len(levels) and levels[i + count] == lv:
            count += 1
        runs.append(f"{'#' * lv}×{count}")
        i += count
    return " ".join(runs)


def _get_file_content_local(repo_path, ref, filepath):
    try:
        return subprocess.check_output(
            ["git", "-C", repo_path, "show", f"{ref}:{filepath}"],
            text=True,
            stderr=subprocess.DEVNULL,
        )
    except subprocess.CalledProcessError:
        return None


def _get_file_content_api(github_client, owner, repo, ref, filepath):
    try:
        repository = github_client.get_repo(f"{owner}/{repo}")
        blob = repository.get_contents(filepath, ref=ref)
        return blob.decoded_content.decode("utf-8")
    except Exception:
        return None


def _source_content_context():
    """Build a context dict for fetching source file content."""
    if MODE == "commit-based":
        owner, repo, _, head = parse_compare_url(SOURCE_COMMIT_COMPARE)
        return {
            "owner": owner,
            "repo": repo,
            "head_ref": head,
            "repo_path": SOURCE_REPO_PATH if os.path.isdir(SOURCE_REPO_PATH or "") else None,
        }
    owner, repo, _ = parse_pr_url(SOURCE_PR)
    return {"owner": owner, "repo": repo, "head_ref": None, "repo_path": None}


def _target_content_context(github_client):
    """Build a context dict for fetching target file content."""
    owner, repo, pr_num = parse_pr_url(TARGET_PR)
    repository = github_client.get_repo(f"{owner}/{repo}")
    pr = repository.get_pull(pr_num)
    return {"owner": owner, "repo": repo, "head_ref": pr.head.sha}


def _fetch_content(ctx, filepath, github_client):
    """Fetch file content using local git or the API."""
    if ctx.get("repo_path"):
        content = _get_file_content_local(ctx["repo_path"], ctx["head_ref"], filepath)
        if content is not None:
            return content
    if ctx.get("head_ref"):
        return _get_file_content_api(
            github_client, ctx["owner"], ctx["repo"], ctx["head_ref"], filepath
        )
    return None


def _structure_worker_count(total, max_workers=None):
    configured_workers = DOCUMENT_STRUCTURE_WORKERS
    if max_workers is not None:
        try:
            configured_workers = max(1, int(max_workers))
        except (TypeError, ValueError):
            configured_workers = DOCUMENT_STRUCTURE_WORKERS
    return min(total, configured_workers)


def _build_heading_structure(filepath, src, tgt, is_added=False):
    src_headings = _extract_headings(src) if src else []
    tgt_headings = _extract_headings(tgt) if tgt else []
    src_levels = [h[0] for h in src_headings]
    tgt_levels = [h[0] for h in tgt_headings]

    match = src_levels == tgt_levels
    line_integrity = None

    if is_added and src and tgt:
        line_integrity = compare_added_file_line_integrity(src, tgt)
        if not line_integrity["heading_lines_match"] or not line_integrity["total_lines_match"]:
            match = False

    return {
        "file": filepath,
        "source_headings": src_headings,
        "target_headings": tgt_headings,
        "source_compact": _compact_structure(src_headings),
        "target_compact": _compact_structure(tgt_headings),
        "match": match,
        "is_added_file": is_added,
        "line_integrity": line_integrity,
    }


def collect_heading_structures(exceed_files, source_ctx, target_ctx, github_client):
    """Compare heading structures of *exceed_files* between source and target.

    Returns a list of dicts, one per file, with heading comparison data.
    """
    results = []
    total = len(exceed_files)
    for idx, filepath in enumerate(exceed_files, 1):
        print(f"    [{idx}/{total}] {filepath}")

        src = _fetch_content(source_ctx, filepath, github_client)
        tgt = _fetch_content(target_ctx, filepath, github_client)

        results.append(_build_heading_structure(filepath, src, tgt))
    return results


def _build_custom_content_structure(filepath, src, tgt):
    src_tags = extract_custom_content_tags(src) if src is not None else []
    tgt_tags = extract_custom_content_tags(tgt) if tgt is not None else []

    if src is None:
        issue_reason = "could not read source HEAD content"
        first_difference = ""
    elif tgt is None:
        issue_reason = "could not read translated target content"
        first_difference = ""
    else:
        issue = compare_custom_content_structure(filepath, src, tgt)
        issue_reason = issue.reason if issue else ""
        first_difference = issue.first_difference if issue else ""

    return {
        "file": filepath,
        "source_tags": src_tags,
        "target_tags": tgt_tags,
        "source_compact": compact_custom_content_tags(src_tags),
        "target_compact": compact_custom_content_tags(tgt_tags),
        "match": not issue_reason,
        "issue": issue_reason,
        "first_difference": first_difference,
    }


def _collect_custom_content_structure(filepath, source_ctx, target_ctx, github_client):
    src = _fetch_content(source_ctx, filepath, github_client)
    tgt = _fetch_content(target_ctx, filepath, github_client)
    return _build_custom_content_structure(filepath, src, tgt)


def _custom_content_failure_row(filepath, exc):
    return {
        "file": filepath,
        "source_tags": [],
        "target_tags": [],
        "source_compact": "0 CustomContent tags",
        "target_compact": "0 CustomContent tags",
        "match": False,
        "issue": f"CustomContent comparison failed: {exc}",
        "first_difference": "",
    }


def _collect_document_structures(
    filepath, source_ctx, target_ctx, github_client, include_heading,
    is_added=False, commit_based_mode=False,
):
    src = _fetch_content(source_ctx, filepath, github_client)
    tgt = _fetch_content(target_ctx, filepath, github_client)

    if commit_based_mode and src:
        src = strip_related_resources_sections(src)

    return {
        "custom_content": _build_custom_content_structure(filepath, src, tgt),
        "heading": (
            _build_heading_structure(filepath, src, tgt, is_added=is_added)
            if include_heading else None
        ),
    }


def collect_custom_content_structures(
    file_paths, source_ctx, target_ctx, github_client, max_workers=None
):
    """Compare CustomContent tag sequences between full source and target docs."""
    file_paths = list(file_paths)
    total = len(file_paths)
    if not total:
        return []

    worker_count = _structure_worker_count(total, max_workers)
    if worker_count <= 1:
        results = []
        for idx, filepath in enumerate(file_paths, 1):
            print(f"    [{idx}/{total}] {filepath}")
            try:
                results.append(
                    _collect_custom_content_structure(filepath, source_ctx, target_ctx, github_client)
                )
            except Exception as exc:
                results.append(_custom_content_failure_row(filepath, exc))
        return results

    print(f"    Using {worker_count} CustomContent worker(s)")
    results = [None] * total
    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        futures = {
            executor.submit(
                _collect_custom_content_structure,
                filepath,
                source_ctx,
                target_ctx,
                github_client,
            ): (index, filepath)
            for index, filepath in enumerate(file_paths)
        }
        for completed, future in enumerate(as_completed(futures), 1):
            index, filepath = futures[future]
            print(f"    [{completed}/{total}] {filepath}")
            try:
                results[index] = future.result()
            except Exception as exc:
                results[index] = _custom_content_failure_row(filepath, exc)

    return results


def collect_document_structures(
    file_paths, heading_files, source_ctx, target_ctx, github_client,
    max_workers=None, source_stats=None, mode="commit-based",
):
    """Collect CustomContent and selected heading structures with one fetch per file."""
    file_paths = list(file_paths)
    heading_files = set(heading_files)
    total = len(file_paths)
    if not total:
        return [], []

    source_stats = source_stats or {}
    commit_based_mode = mode == "commit-based"
    added_files = {
        fp for fp in file_paths
        if source_stats.get(fp, {}).get("status") == "added"
    }

    worker_count = _structure_worker_count(total, max_workers)
    custom_results = [None] * total
    heading_results_by_index = {}

    if worker_count <= 1:
        for index, filepath in enumerate(file_paths):
            print(f"    [{index + 1}/{total}] {filepath}")
            include_heading = filepath in heading_files
            try:
                result = _collect_document_structures(
                    filepath, source_ctx, target_ctx, github_client, include_heading,
                    is_added=filepath in added_files,
                    commit_based_mode=commit_based_mode,
                )
            except Exception as exc:
                result = {
                    "custom_content": _custom_content_failure_row(filepath, exc),
                    "heading": None,
                }
            custom_results[index] = result["custom_content"]
            if include_heading and result["heading"] is not None:
                heading_results_by_index[index] = result["heading"]
        heading_results = [
            heading_results_by_index[i] for i in sorted(heading_results_by_index)
        ]
        return custom_results, heading_results

    print(f"    Using {worker_count} document structure worker(s)")
    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        futures = {
            executor.submit(
                _collect_document_structures,
                filepath,
                source_ctx,
                target_ctx,
                github_client,
                filepath in heading_files,
                filepath in added_files,
                commit_based_mode,
            ): (index, filepath)
            for index, filepath in enumerate(file_paths)
        }
        for completed, future in enumerate(as_completed(futures), 1):
            index, filepath = futures[future]
            print(f"    [{completed}/{total}] {filepath}")
            include_heading = filepath in heading_files
            try:
                result = future.result()
            except Exception as exc:
                result = {
                    "custom_content": _custom_content_failure_row(filepath, exc),
                    "heading": None,
                }
            custom_results[index] = result["custom_content"]
            if include_heading and result["heading"] is not None:
                heading_results_by_index[index] = result["heading"]

    heading_results = [
        heading_results_by_index[i] for i in sorted(heading_results_by_index)
    ]
    return custom_results, heading_results


def has_custom_content_report_entry(data):
    """Return True when a CustomContent row should appear in the report."""
    return bool(data.get("source_tags") or data.get("target_tags") or data.get("issue"))


# ── Report Building ──────────────────────────────────────────────────────────

def build_report_rows(source_stats, target_stats, threshold):
    """
    Compare target files against source and return a list of row dicts.
    Only files present in the target are included.

    Each md row gets a ``level`` field:
      - "exact"     – additions AND deletions match exactly
      - "within"    – diff exists but within *threshold*
      - "exceed"    – diff exceeds *threshold*
      - "target_only" – file only in target, not in source diff
    Non-md rows get level = "".
    """
    rows = []
    for filepath in sorted(target_stats):
        basename = os.path.basename(filepath)
        if basename in EXCLUDE_FILES:
            continue

        t = target_stats[filepath]
        s = source_stats.get(filepath)

        row = {
            "file": filepath,
            "is_md": t["is_md"],
            "target_status": t["status"],
            "target_additions": t["additions"],
            "target_deletions": t["deletions"],
            "in_source": s is not None,
        }

        if s:
            row["source_status"] = s["status"]
            row["source_additions"] = s["additions"]
            row["source_deletions"] = s["deletions"]
        else:
            row["source_status"] = "—"
            row["source_additions"] = 0
            row["source_deletions"] = 0

        if t["is_md"]:
            add_diff = abs(t["additions"] - (s["additions"] if s else 0))
            del_diff = abs(t["deletions"] - (s["deletions"] if s else 0))
            row["add_diff"] = add_diff
            row["del_diff"] = del_diff
            if not s:
                row["level"] = "target_only"
            elif add_diff == 0 and del_diff == 0:
                row["level"] = "exact"
            elif add_diff > threshold or del_diff > threshold:
                row["level"] = "exceed"
            else:
                row["level"] = "within"
        else:
            row["add_diff"] = ""
            row["del_diff"] = ""
            row["level"] = ""

        rows.append(row)

    return rows


# ── Excel Output ─────────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_EXACT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # green
_WITHIN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # yellow
_EXCEED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # red
_TARGET_ONLY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # gray
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_LEVEL_FILL = {
    "exact": _EXACT_FILL,
    "within": _WITHIN_FILL,
    "exceed": _EXCEED_FILL,
    "target_only": _TARGET_ONLY_FILL,
}


def _write_detail_sheet(wb, rows, threshold):
    """Write the main detail sheet with per-file comparison."""
    ws = wb.active
    ws.title = "Structure Details"

    headers = [
        "File",
        "Type",
        "In Source",
        "Source Status",
        "Target Status",
        "Source +",
        "Source −",
        "Target +",
        "Target −",
        "|Δ Add|",
        "|Δ Del|",
        "Status",
        "Notes",
    ]
    col_widths = [60, 6, 10, 13, 13, 10, 10, 10, 10, 9, 9, 15, 32]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    ws.row_dimensions[1].height = 32

    status_labels = {
        "exact": "Exact match",
        "within": "Within threshold",
        "exceed": "Exceeds threshold",
        "target_only": "Target-only",
    }

    for i, row in enumerate(rows, 2):
        level = row.get("level", "")
        values = [
            row["file"],
            "md" if row["is_md"] else os.path.splitext(row["file"])[1].lstrip(".") or "other",
            "Yes" if row["in_source"] else "No",
            row["source_status"],
            row["target_status"],
            row["source_additions"],
            row["source_deletions"],
            row["target_additions"],
            row["target_deletions"],
            row["add_diff"],
            row["del_diff"],
            status_labels.get(level, ""),
            "" if row["in_source"] else "Not in source diff",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(
                horizontal="left" if col == 1 else "center",
                vertical="center",
            )

        fill = _LEVEL_FILL.get(level)
        if fill:
            for col in range(1, len(values) + 1):
                ws.cell(row=i, column=col).fill = fill

    for idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    ws.freeze_panes = "A2"


def _write_summary_sheet(wb, rows, source_label, target_label, threshold):
    """Write a summary sheet with aggregate statistics."""
    ws = wb.create_sheet("Summary", 0)  # insert before Detail

    md_rows = [r for r in rows if r["is_md"]]
    non_md_rows = [r for r in rows if not r["is_md"]]
    md_exact = [r for r in md_rows if r.get("level") == "exact"]
    md_within = [r for r in md_rows if r.get("level") == "within"]
    md_exceed = [r for r in md_rows if r.get("level") == "exceed"]
    md_target_only = [r for r in md_rows if r.get("level") == "target_only"]

    title_font = Font(bold=True, size=14, color="1F4E79")
    section_font = Font(bold=True, size=12, color="2E75B6")
    label_font = Font(size=11)
    value_font = Font(size=11, bold=True)
    good_font = Font(size=11, bold=True, color="006100")
    bad_font = Font(size=11, bold=True, color="9C0006")

    row_num = 1

    def _add_row(label, value, label_f=label_font, value_f=value_font, indent=2):
        nonlocal row_num
        ws.cell(row=row_num, column=indent, value=label).font = label_f
        ws.cell(row=row_num, column=indent + 2, value=value).font = value_f
        row_num += 1

    ws.cell(row=row_num, column=1, value="Translation Verification Report").font = title_font
    row_num += 2

    ws.cell(row=row_num, column=1, value="Configuration").font = section_font
    row_num += 1
    _add_row("Source:", source_label)
    _add_row("Target:", target_label)
    _add_row("Mode:", MODE)
    _add_row("Warning threshold:", f">{threshold} lines")
    _add_row("Generated at:", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    row_num += 1

    ws.cell(row=row_num, column=1, value="Markdown Files").font = section_font
    row_num += 1
    _add_row("Total md files in target:", len(md_rows))
    _add_row("  Exact match:", len(md_exact), value_f=good_font)
    _add_row("  Within threshold:", len(md_within), value_f=value_font)
    _add_row("  Exceeds threshold:", len(md_exceed), value_f=bad_font)
    _add_row("  Target-only:", len(md_target_only))
    row_num += 1

    if non_md_rows:
        ws.cell(row=row_num, column=1, value="Non-Markdown Files").font = section_font
        row_num += 1
        _add_row("Total non-md files in target:", len(non_md_rows))
        row_num += 1

    ws.cell(row=row_num, column=1, value="Legend (Detail sheet)").font = section_font
    row_num += 1
    legends = [
        (_EXACT_FILL, "Green — additions & deletions match exactly"),
        (_WITHIN_FILL, "Yellow — diff exists but within threshold"),
        (_EXCEED_FILL, "Red — diff exceeds threshold"),
        (_TARGET_ONLY_FILL, "Gray — file in target only (not in source diff)"),
    ]
    for fill, desc in legends:
        cell = ws.cell(row=row_num, column=2, value="    ")
        cell.fill = fill
        cell.border = _THIN_BORDER
        ws.cell(row=row_num, column=3, value=desc).font = label_font
        row_num += 1

    attention_rows = [r for r in md_rows if r.get("level") in ("exceed", "target_only")]
    if attention_rows:
        row_num += 1
        ws.cell(row=row_num, column=1, value="Files Needing Attention").font = section_font
        row_num += 1
        attn_headers = ["File", "Status", "Source +", "Source −", "Target +", "Target −", "|Δ Add|", "|Δ Del|", "Notes"]
        for col, h in enumerate(attn_headers, 1):
            cell = ws.cell(row=row_num, column=col, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = _THIN_BORDER
        row_num += 1
        status_labels = {
            "exceed": "Exceeds threshold",
            "target_only": "Target-only",
        }
        for r in attention_rows:
            level = r.get("level", "")
            vals = [
                r["file"],
                status_labels.get(level, level),
                r["source_additions"],
                r["source_deletions"],
                r["target_additions"],
                r["target_deletions"],
                r["add_diff"],
                r["del_diff"],
                "" if r["in_source"] else "Not in source diff",
            ]
            fill = _LEVEL_FILL.get(level, _EXCEED_FILL)
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row_num, column=col, value=v)
                cell.fill = fill
                cell.border = _THIN_BORDER
                cell.alignment = Alignment(horizontal="left" if col == 1 else "center")
            row_num += 1

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 52
    for letter in "EFGHIJ":
        ws.column_dimensions[letter].width = 12


_HEADING_DETAIL_HEADERS = ["#", "Source Level", "Source Heading", "Target Level", "Target Heading", "Level Match"]
_HEADING_COL_WIDTHS = [5, 14, 55, 14, 55, 13]
_FILE_FONT = Font(bold=True, size=11, color="1F4E79")
_FILE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_MATCH_FONT = Font(bold=True, color="006100")
_MISMATCH_FONT = Font(bold=True, color="9C0006")
_MISSING_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def _write_heading_detail_rows(ws, row_num, file_list):
    """Write heading-by-heading detail rows for each file in *file_list*.

    Returns the next available row number.
    """
    headers = _HEADING_DETAIL_HEADERS
    for hd in file_list:
        row_num += 1
        file_cell = ws.cell(row=row_num, column=1, value=hd["file"])
        file_cell.font = _FILE_FONT
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).fill = _FILE_FILL
            ws.cell(row=row_num, column=col).border = _THIN_BORDER
        ws.cell(
            row=row_num, column=len(headers),
            value=f"Source: {len(hd['source_headings'])} headings  |  Target: {len(hd['target_headings'])} headings",
        ).font = Font(italic=True, size=10, color="404040")
        row_num += 1

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = _THIN_BORDER
        row_num += 1

        src_h = hd["source_headings"]
        tgt_h = hd["target_headings"]
        max_len = max(len(src_h), len(tgt_h)) if (src_h or tgt_h) else 0

        for pos in range(max_len):
            s_lv = src_h[pos][0] if pos < len(src_h) else None
            s_txt = src_h[pos][1] if pos < len(src_h) else ""
            t_lv = tgt_h[pos][0] if pos < len(tgt_h) else None
            t_txt = tgt_h[pos][1] if pos < len(tgt_h) else ""

            s_lv_str = "#" * s_lv if s_lv else "(missing)"
            t_lv_str = "#" * t_lv if t_lv else "(missing)"
            levels_match = s_lv is not None and t_lv is not None and s_lv == t_lv

            vals = [pos + 1, s_lv_str, s_txt, t_lv_str, t_txt]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row_num, column=col, value=v)
                cell.border = _THIN_BORDER
                cell.alignment = Alignment(
                    horizontal="center" if col in (1, 2, 4) else "left",
                    vertical="center",
                )

            match_cell = ws.cell(row=row_num, column=6)
            match_cell.border = _THIN_BORDER
            match_cell.alignment = Alignment(horizontal="center")

            if levels_match:
                match_cell.value = "✓"
                match_cell.font = _MATCH_FONT
                fill = _EXACT_FILL
            else:
                match_cell.value = "✗"
                match_cell.font = _MISMATCH_FONT
                fill = _EXCEED_FILL
                if s_lv is None or t_lv is None:
                    fill = _MISSING_FILL

            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = fill
            row_num += 1

    return row_num


def _apply_heading_col_widths(ws):
    for i, w in enumerate(_HEADING_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(
            ws.column_dimensions[get_column_letter(i)].width or 0, w
        )


def _write_line_integrity_detail_rows(ws, row_num, file_list):
    """Write heading-line-position and total-line-count detail for added files."""
    headers = ["#", "Source Line", "Source Heading", "Target Line", "Target Heading", "Line Match"]
    for hd in file_list:
        li = hd["line_integrity"]

        row_num += 1
        file_cell = ws.cell(row=row_num, column=1, value=hd["file"])
        file_cell.font = _FILE_FONT
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).fill = _FILE_FILL
            ws.cell(row=row_num, column=col).border = _THIN_BORDER
        total_note = (
            f"Source: {li['source_total_lines']} lines  |  "
            f"Target: {li['target_total_lines']} lines"
        )
        if not li["total_lines_match"]:
            total_note += "  ← TOTAL LINE COUNT MISMATCH"
        ws.cell(row=row_num, column=len(headers), value=total_note).font = Font(
            italic=True, size=10, color="9C0006" if not li["total_lines_match"] else "404040",
        )
        row_num += 1

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = _THIN_BORDER
        row_num += 1

        for mh in li["mismatched_headings"]:
            vals = [
                mh["index"],
                mh["source_line"] if mh["source_line"] is not None else "(missing)",
                mh["source_heading"],
                mh["target_line"] if mh["target_line"] is not None else "(missing)",
                mh["target_heading"],
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row_num, column=col, value=v)
                cell.border = _THIN_BORDER
                cell.alignment = Alignment(
                    horizontal="center" if col in (1, 2, 4) else "left",
                    vertical="center",
                )

            match_cell = ws.cell(row=row_num, column=6)
            match_cell.border = _THIN_BORDER
            match_cell.alignment = Alignment(horizontal="center")
            match_cell.value = "✗"
            match_cell.font = _MISMATCH_FONT
            fill = _MISSING_FILL if mh["source_line"] is None or mh["target_line"] is None else _EXCEED_FILL
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = fill
            row_num += 1

        if not li["mismatched_headings"] and not li["total_lines_match"]:
            ws.cell(
                row=row_num, column=1,
                value="All heading line positions match, but total line count differs.",
            ).font = Font(italic=True, size=10, color="9C0006")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).border = _THIN_BORDER
                ws.cell(row=row_num, column=col).fill = _WITHIN_FILL
            row_num += 1

    return row_num


def _write_heading_sheet(wb, heading_data):
    """Write the main heading-structure sheet (overview + mismatched detail)."""
    if not heading_data:
        return
    ws = wb.create_sheet("Heading Structure")

    matched_list = [hd for hd in heading_data if hd["match"]]
    mismatched_list = [hd for hd in heading_data if not hd["match"]]

    row_num = 1
    ws.cell(row=row_num, column=1, value="Heading Structure Comparison").font = Font(
        bold=True, size=14, color="1F4E79"
    )
    row_num += 1
    ws.cell(
        row=row_num, column=1,
        value=(
            f"Files analyzed: {len(heading_data)}  |  "
            f"Structure match: {len(matched_list)}  |  "
            f"Structure mismatch: {len(mismatched_list)}"
        ),
    ).font = Font(size=11, color="404040")
    row_num += 2

    # ── Overview table ────────────────────────────────────────────────────
    has_added = any(hd.get("is_added_file") for hd in heading_data)
    overview_headers = [
        "File", "Source Headings", "Target Headings",
        "Source Structure", "Target Structure", "Match",
    ]
    if has_added:
        overview_headers += ["Added File", "Src Lines", "Tgt Lines", "Line Integrity"]
    for col, h in enumerate(overview_headers, 1):
        cell = ws.cell(row=row_num, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    row_num += 1

    for hd in mismatched_list + matched_list:
        li = hd.get("line_integrity")
        vals = [
            hd["file"],
            len(hd["source_headings"]),
            len(hd["target_headings"]),
            hd["source_compact"],
            hd["target_compact"],
            "Match" if hd["match"] else "Mismatch",
        ]
        if has_added:
            if hd.get("is_added_file") and li:
                li_ok = li["heading_lines_match"] and li["total_lines_match"]
                vals += [
                    "Yes",
                    li["source_total_lines"],
                    li["target_total_lines"],
                    "OK" if li_ok else "Mismatch",
                ]
            else:
                vals += ["No", "", "", "—"]
        fill = _EXACT_FILL if hd["match"] else _EXCEED_FILL
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.fill = fill
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(
                horizontal="left" if col in (1, 4, 5) else "center",
                vertical="center",
            )
        row_num += 1

    overview_widths = [60, 16, 16, 50, 50, 12]
    if has_added:
        overview_widths += [12, 10, 10, 14]
    for i, w in enumerate(overview_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(
            ws.column_dimensions[get_column_letter(i)].width or 0, w
        )

    # ── Detailed comparison for mismatched files ──────────────────────────
    if mismatched_list:
        row_num += 2
        ws.cell(
            row=row_num, column=1,
            value=f"Detailed Heading Comparison — Structure Mismatch ({len(mismatched_list)} files)",
        ).font = Font(bold=True, size=12, color="9C0006")
        row_num += 1
        row_num = _write_heading_detail_rows(ws, row_num, mismatched_list)

    # ── Line integrity detail for added files with mismatches ─────────────
    li_mismatch_files = [
        hd for hd in heading_data
        if hd.get("is_added_file") and hd.get("line_integrity")
        and (
            not hd["line_integrity"]["heading_lines_match"]
            or not hd["line_integrity"]["total_lines_match"]
        )
    ]
    if li_mismatch_files:
        row_num += 2
        ws.cell(
            row=row_num, column=1,
            value=f"Added-File Line Integrity — Mismatch ({len(li_mismatch_files)} files)",
        ).font = Font(bold=True, size=12, color="9C0006")
        row_num += 1
        row_num = _write_line_integrity_detail_rows(ws, row_num, li_mismatch_files)

    _apply_heading_col_widths(ws)


def _write_heading_matched_sheet(wb, heading_data):
    """Write a reference sheet with heading detail for structure-matched files."""
    matched_list = [hd for hd in heading_data if hd["match"]]
    if not matched_list:
        return
    ws = wb.create_sheet("Headings (Matched)")

    row_num = 1
    ws.cell(
        row=row_num, column=1,
        value=f"Heading Detail — Structure Match ({len(matched_list)} files, for reference)",
    ).font = Font(bold=True, size=14, color="1F4E79")
    row_num += 1
    ws.cell(
        row=row_num, column=1,
        value="These files exceed the line-count threshold but have identical heading structure.",
    ).font = Font(size=11, color="404040")
    row_num += 1

    row_num = _write_heading_detail_rows(ws, row_num, matched_list)
    _apply_heading_col_widths(ws)


_CUSTOM_CONTENT_DETAIL_HEADERS = ["#", "Source Line", "Source Tag", "Target Line", "Target Tag", "Tag Match"]
_CUSTOM_CONTENT_COL_WIDTHS = [5, 12, 60, 12, 60, 12]


def _tag_signature(tag):
    return (tag.kind, tag.text) if tag else None


def _tag_line(tag):
    return tag.line_number if tag else ""


def _tag_text(tag):
    return tag.text if tag else "(missing)"


def _write_custom_content_detail_rows(ws, row_num, file_list):
    """Write tag-by-tag CustomContent detail rows for mismatched files."""
    headers = _CUSTOM_CONTENT_DETAIL_HEADERS
    for data in file_list:
        row_num += 1
        file_cell = ws.cell(row=row_num, column=1, value=data["file"])
        file_cell.font = _FILE_FONT
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).fill = _FILE_FILL
            ws.cell(row=row_num, column=col).border = _THIN_BORDER
        ws.cell(
            row=row_num, column=len(headers),
            value=(
                f"Source: {len(data['source_tags'])} tags  |  "
                f"Target: {len(data['target_tags'])} tags"
            ),
        ).font = Font(italic=True, size=10, color="404040")
        row_num += 1

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = _THIN_BORDER
        row_num += 1

        source_tags = data["source_tags"]
        target_tags = data["target_tags"]
        max_len = max(len(source_tags), len(target_tags)) if (source_tags or target_tags) else 0

        for pos in range(max_len):
            src_tag = source_tags[pos] if pos < len(source_tags) else None
            tgt_tag = target_tags[pos] if pos < len(target_tags) else None
            tags_match = _tag_signature(src_tag) == _tag_signature(tgt_tag)

            values = [
                pos + 1,
                _tag_line(src_tag),
                _tag_text(src_tag),
                _tag_line(tgt_tag),
                _tag_text(tgt_tag),
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = _THIN_BORDER
                cell.alignment = Alignment(
                    horizontal="center" if col in (1, 2, 4) else "left",
                    vertical="center",
                )

            match_cell = ws.cell(row=row_num, column=6)
            match_cell.border = _THIN_BORDER
            match_cell.alignment = Alignment(horizontal="center")
            if tags_match:
                match_cell.value = "✓"
                match_cell.font = _MATCH_FONT
                fill = _EXACT_FILL
            else:
                match_cell.value = "✗"
                match_cell.font = _MISMATCH_FONT
                fill = _MISSING_FILL if src_tag is None or tgt_tag is None else _EXCEED_FILL

            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = fill
            row_num += 1

    return row_num


def _write_custom_content_sheet(wb, custom_content_data):
    """Write CustomContent structure comparison sheet."""
    report_data = [
        data for data in (custom_content_data or [])
        if has_custom_content_report_entry(data)
    ]
    if not report_data:
        return

    ws = wb.create_sheet("CustomContent")
    matched_list = [data for data in report_data if data["match"]]
    mismatched_list = [data for data in report_data if not data["match"]]

    row_num = 1
    ws.cell(row=row_num, column=1, value="CustomContent Structure Comparison").font = Font(
        bold=True, size=14, color="1F4E79"
    )
    row_num += 1
    ws.cell(
        row=row_num, column=1,
        value=(
            f"Files with CustomContent: {len(report_data)}  |  "
            f"Match: {len(matched_list)}  |  "
            f"Mismatch: {len(mismatched_list)}"
        ),
    ).font = Font(size=11, color="404040")
    row_num += 2

    overview_headers = [
        "File",
        "Source Tags",
        "Target Tags",
        "Source CustomContent",
        "Target CustomContent",
        "Result",
        "Issue",
        "First Difference",
    ]
    for col, header in enumerate(overview_headers, 1):
        cell = ws.cell(row=row_num, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    row_num += 1

    for data in mismatched_list + matched_list:
        values = [
            data["file"],
            len(data["source_tags"]),
            len(data["target_tags"]),
            data["source_compact"],
            data["target_compact"],
            "Match" if data["match"] else "Mismatch",
            data["issue"],
            data["first_difference"],
        ]
        fill = _EXACT_FILL if data["match"] else _EXCEED_FILL
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.fill = fill
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(
                horizontal="left" if col in (1, 4, 5, 7, 8) else "center",
                vertical="center",
                wrap_text=col in (4, 5, 7, 8),
            )
        row_num += 1

    overview_widths = [60, 13, 13, 70, 70, 12, 34, 70]
    for idx, width in enumerate(overview_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    if mismatched_list:
        row_num += 2
        ws.cell(
            row=row_num, column=1,
            value=f"Detailed CustomContent Comparison — Mismatch ({len(mismatched_list)} files)",
        ).font = Font(bold=True, size=12, color="9C0006")
        row_num += 1
        row_num = _write_custom_content_detail_rows(ws, row_num, mismatched_list)

    for idx, width in enumerate(_CUSTOM_CONTENT_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = max(
            ws.column_dimensions[get_column_letter(idx)].width or 0,
            width,
        )


def write_excel(rows, output_path, source_label, target_label, threshold, heading_data=None, custom_content_data=None):
    """Generate the full Excel report."""
    wb = Workbook()
    _write_detail_sheet(wb, rows, threshold)
    _write_summary_sheet(wb, rows, source_label, target_label, threshold)
    if custom_content_data:
        _write_custom_content_sheet(wb, custom_content_data)
    if heading_data:
        _write_heading_sheet(wb, heading_data)
        _write_heading_matched_sheet(wb, heading_data)
    wb.save(output_path)
    print(f"\n  Report saved to: {output_path}")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    if not GITHUB_TOKEN:
        print("Error: GITHUB_TOKEN environment variable is required.")
        print("  export GITHUB_TOKEN='ghp_...'")
        sys.exit(1)

    github_client = Github(auth=Auth.Token(GITHUB_TOKEN))

    print("\n[1/4] Fetching source file stats...")
    source_stats, source_label = get_source_stats(github_client)

    print("\n[2/4] Fetching target file stats...")
    target_stats, target_label = get_target_stats(github_client)

    print("\n[3/4] Building comparison report...")
    rows = build_report_rows(source_stats, target_stats, WARNING_THRESHOLD)

    md_rows = [r for r in rows if r["is_md"]]
    md_exact = [r for r in md_rows if r.get("level") == "exact"]
    md_within = [r for r in md_rows if r.get("level") == "within"]
    md_exceed = [r for r in md_rows if r.get("level") == "exceed"]
    md_target_only = [r for r in md_rows if r.get("level") == "target_only"]

    print(f"\n  Summary:")
    print(f"    Source files:        {len(source_stats)}")
    print(f"    Target files:        {len(target_stats)}")
    print(f"    MD files in target:  {len(md_rows)}")
    print(f"    Exact match:         {len(md_exact)}")
    print(f"    Within threshold:    {len(md_within)}")
    print(f"    Exceeds threshold:   {len(md_exceed)}")
    print(f"    Target-only:         {len(md_target_only)}")

    source_ctx = _source_content_context()
    target_ctx = _target_content_context(github_client)

    custom_content_data = []
    custom_mismatch = 0
    heading_data = []
    if md_rows:
        print(f"\n[4/4] Comparing document structures for {len(md_rows)} changed markdown files...")
        md_files = [r["file"] for r in md_rows]
        exceed_files = [r["file"] for r in md_exceed]
        added_files = [
            r["file"] for r in md_rows
            if source_stats.get(r["file"], {}).get("status") == "added"
        ]
        heading_files = list(set(exceed_files) | set(added_files))
        custom_content_data, heading_data = collect_document_structures(
            md_files, heading_files, source_ctx, target_ctx, github_client,
            source_stats=source_stats, mode=MODE,
        )
        custom_report_data = [
            item for item in custom_content_data
            if has_custom_content_report_entry(item)
        ]
        custom_match = sum(1 for item in custom_report_data if item["match"])
        custom_mismatch = len(custom_report_data) - custom_match
        custom_skipped = len(custom_content_data) - len(custom_report_data)
        print(
            f"  CustomContent structure: {custom_match} match, "
            f"{custom_mismatch} mismatch, {custom_skipped} without CustomContent"
        )
        if heading_files:
            h_match = sum(1 for h in heading_data if h["match"])
            h_mismatch = len(heading_data) - h_match
            print(f"  Heading structure: {h_match} match, {h_mismatch} mismatch")
            li_files = [h for h in heading_data if h.get("line_integrity")]
            if li_files:
                li_ok = sum(
                    1 for h in li_files
                    if h["line_integrity"]["heading_lines_match"]
                    and h["line_integrity"]["total_lines_match"]
                )
                print(f"  Added-file line integrity: {li_ok}/{len(li_files)} pass")
        else:
            print("  Heading structure: no exceed-threshold or added files — skipped")
    else:
        print("\n[4/4] No markdown files — skipping document structure analysis.")

    script_dir = os.path.dirname(os.path.abspath(__file__))
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(script_dir, f"translation_verify_{timestamp}.xlsx")

    write_excel(
        rows,
        output_path,
        source_label,
        target_label,
        WARNING_THRESHOLD,
        heading_data,
        custom_content_data,
    )

    line_count_flagged = len(md_exceed) + len(md_target_only)
    flagged = line_count_flagged + custom_mismatch
    if flagged:
        print(
            f"\n  ⚠  {line_count_flagged} line-count file(s) and "
            f"{custom_mismatch} CustomContent file(s) need attention — review the Excel for details."
        )
    else:
        print(f"\n  ✓  All md files match or are within threshold, and CustomContent tags match.")


if __name__ == "__main__":
    main()
