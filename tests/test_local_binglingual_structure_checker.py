import importlib.util
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


SCRIPTS_DIR = Path(__file__).resolve().parents[1] / "scripts"
sys.path.insert(0, str(SCRIPTS_DIR))

SPEC = importlib.util.spec_from_file_location(
    "local_binglingual_structure_checker",
    SCRIPTS_DIR / "local-binglingual-structure-checker.py",
)
checker = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(checker)


class LocalBinglingualStructureCheckerTest(unittest.TestCase):
    def test_check_file_reports_heading_and_custom_content_issues(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            source_root = root / "source"
            target_root = root / "target"
            source_root.mkdir()
            target_root.mkdir()

            rel_path = "guide.md"
            (source_root / rel_path).write_text(
                '<CustomContent plan="essential">\n\n# Guide\n\n## Section\n\n</CustomContent>\n',
                encoding="utf-8",
            )
            (target_root / rel_path).write_text(
                "# 指南\n\n### Section\n",
                encoding="utf-8",
            )

            issues = checker.check_file(source_root, target_root, rel_path)

        self.assertEqual(
            ["heading level sequence differs", "CustomContent tag sequence differs"],
            [issue.reason for issue in issues],
        )

    def test_check_file_reports_missing_target(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            source_root = root / "source"
            target_root = root / "target"
            source_root.mkdir()
            target_root.mkdir()
            (source_root / "guide.md").write_text("# Guide\n", encoding="utf-8")

            issues = checker.check_file(source_root, target_root, "guide.md")

        self.assertEqual(["target file missing"], [issue.reason for issue in issues])

    def test_write_excel_report_splits_issue_types_into_sheets(self):
        issues = [
            checker.StructureValidationIssue(
                "heading.md",
                "heading level sequence differs",
                source_compact="#x1 ##x1",
                target_compact="#x1 ###x1",
                first_difference="heading 2: source ##, target ###",
            ),
            checker.StructureValidationIssue(
                "custom.md",
                "CustomContent tag sequence differs",
                source_compact=(
                    '2 CustomContent tags: <CustomContent plan="essential"> | '
                    "</CustomContent>"
                ),
                target_compact="0 CustomContent tags",
                first_difference=(
                    'CustomContent tag 1: source <CustomContent plan="essential"> '
                    "at line 1, target (missing)"
                ),
            ),
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            source_root = root / "source"
            target_root = root / "target"
            source_root.mkdir()
            target_root.mkdir()
            (source_root / "heading.md").write_text(
                "# Title\n\n## Step\n",
                encoding="utf-8",
            )
            (target_root / "heading.md").write_text(
                "# タイトル\n\n### ステップ\n",
                encoding="utf-8",
            )
            (source_root / "custom.md").write_text(
                '<CustomContent plan="essential">\n\n# Guide\n\n</CustomContent>\n',
                encoding="utf-8",
            )
            (target_root / "custom.md").write_text(
                "# ガイド\n",
                encoding="utf-8",
            )

            output = root / "issues.xlsx"
            checker.write_excel_report(
                issues, output, source_root, target_root,
                source_label="English", target_label="Japanese",
            )
            workbook = load_workbook(output)
            heading_rows = list(workbook["Heading"].iter_rows(values_only=True))
            custom_rows = list(workbook["CustomContent"].iter_rows(values_only=True))

        self.assertEqual(["Heading", "CustomContent"], workbook.sheetnames)
        self.assertEqual(
            (
                "#",
                "English Line",
                "English Level",
                "English Heading",
                "Japanese Line",
                "Japanese Level",
                "Japanese Heading",
                "Level Match",
            ),
            heading_rows[1],
        )
        self.assertEqual(
            (
                "#",
                "English Line",
                "English Tag",
                "Japanese Line",
                "Japanese Tag",
                "Tag Match",
            ),
            custom_rows[1],
        )
        self.assertEqual("heading.md", heading_rows[0][0])
        self.assertEqual("heading level sequence differs", heading_rows[0][1])
        self.assertEqual((1, 1, "#", "Title", 1, "#", "タイトル", "Match"), heading_rows[2])
        self.assertEqual((2, 3, "##", "Step", 3, "###", "ステップ", "Mismatch"), heading_rows[3])
        self.assertEqual("custom.md", custom_rows[0][0])
        self.assertEqual("CustomContent tag sequence differs", custom_rows[0][1])
        self.assertEqual(
            (1, 1, '<CustomContent plan="essential">', None, "(missing)", "Mismatch"),
            custom_rows[2],
        )
        self.assertEqual(
            (2, 5, "</CustomContent>", None, "(missing)", "Mismatch"),
            custom_rows[3],
        )

    def test_discover_all_tocs(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "TOC.md").write_text("- [Guide](/guide.md)\n", encoding="utf-8")
            (root / "TOC-cloud.md").write_text("- [Cloud](/cloud.md)\n", encoding="utf-8")
            (root / "README.md").write_text("# Docs\n", encoding="utf-8")

            tocs = checker.discover_all_tocs(root)

        self.assertEqual(["TOC-cloud.md", "TOC.md"], tocs)


if __name__ == "__main__":
    unittest.main()
