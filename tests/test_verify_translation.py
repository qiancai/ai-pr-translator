import os
import sys
import tempfile
import unittest
from unittest import mock
from pathlib import Path

from openpyxl import load_workbook


SCRIPTS_DIR = Path(__file__).resolve().parents[1] / "scripts"
sys.path.insert(0, str(SCRIPTS_DIR))

from translation_structure_validator import compact_custom_content_tags, extract_custom_content_tags
import verify_translation
from verify_translation import write_excel


def _row(file_path):
    return {
        "file": file_path,
        "is_md": True,
        "target_status": "modified",
        "target_additions": 1,
        "target_deletions": 1,
        "in_source": True,
        "source_status": "modified",
        "source_additions": 1,
        "source_deletions": 1,
        "add_diff": 0,
        "del_diff": 0,
        "level": "exact",
    }


def _custom_content_data(file_path, source_content, target_content, match=True, issue=""):
    source_tags = extract_custom_content_tags(source_content)
    target_tags = extract_custom_content_tags(target_content)
    return {
        "file": file_path,
        "source_tags": source_tags,
        "target_tags": target_tags,
        "source_compact": compact_custom_content_tags(source_tags),
        "target_compact": compact_custom_content_tags(target_tags),
        "match": match,
        "issue": issue,
        "first_difference": "",
    }


class VerifyTranslationReportTest(unittest.TestCase):
    def test_cli_configuration_has_no_repository_specific_defaults(self):
        with mock.patch.dict(os.environ, {}, clear=True):
            args = verify_translation._parse_args(
                [
                    "--source-pr",
                    "https://github.com/acme/docs/pull/1",
                    "--target-pr",
                    "https://github.com/acme/docs-cn/pull/2",
                ]
            )

        verify_translation._apply_runtime_config(args)

        self.assertEqual("pr", verify_translation.MODE)
        self.assertEqual("", verify_translation.SOURCE_COMMIT_COMPARE)
        self.assertEqual("https://github.com/acme/docs/pull/1", verify_translation.SOURCE_PR)
        self.assertEqual("https://github.com/acme/docs-cn/pull/2", verify_translation.TARGET_PR)
        self.assertEqual("", verify_translation.SOURCE_REPO_PATH)

    def test_noop_line_change_pairs_ignore_eof_newline_noise(self):
        patch = "\n".join(
            [
                "@@ -220,2 +220,2 @@",
                '-    <td style="text-align:center;">❌</td>',
                '-    <td style="text-align:center;">❌</td>',
                '+    <td style="text-align:center;">✅</td>',
                '+    <td style="text-align:center;">✅</td>',
                "@@ -385 +385 @@",
                '-> To request a feature in private preview, click **?**.',
                "\\ No newline at end of file",
                '+> To request a feature in private preview, click **?**.',
            ]
        )
        stats = {
            "tidb-cloud/features.md": {
                "status": "modified",
                "additions": 3,
                "deletions": 3,
                "changes": 6,
                "is_md": True,
            }
        }

        verify_translation._apply_noop_line_change_adjustments(
            stats, {"tidb-cloud/features.md": patch}
        )

        self.assertEqual(2, stats["tidb-cloud/features.md"]["additions"])
        self.assertEqual(2, stats["tidb-cloud/features.md"]["deletions"])
        self.assertEqual(4, stats["tidb-cloud/features.md"]["changes"])

    def test_collect_custom_content_structures_parallel_preserves_input_order(self):
        contents = {
            ("source", "wrapped.md"): '<CustomContent plan="essential">\n</CustomContent>\n',
            ("target", "wrapped.md"): '<CustomContent plan="essential">\n</CustomContent>\n',
            ("source", "plain.md"): "# Plain\n",
            ("target", "plain.md"): "# 普通\n",
        }

        def fake_fetch(ctx, filepath, github_client):
            return contents[(ctx["side"], filepath)]

        with mock.patch.object(verify_translation, "_fetch_content", side_effect=fake_fetch) as fetch:
            result = verify_translation.collect_custom_content_structures(
                ["wrapped.md", "plain.md"],
                {"side": "source"},
                {"side": "target"},
                object(),
                max_workers=2,
            )

        self.assertEqual(["wrapped.md", "plain.md"], [item["file"] for item in result])
        self.assertTrue(result[0]["match"])
        self.assertTrue(result[1]["match"])
        self.assertEqual(4, fetch.call_count)

    def test_collect_custom_content_structures_records_fetch_failures(self):
        def fake_fetch(ctx, filepath, github_client):
            raise RuntimeError(f"failed to fetch {filepath}")

        with mock.patch.object(verify_translation, "_fetch_content", side_effect=fake_fetch):
            result = verify_translation.collect_custom_content_structures(
                ["broken.md"],
                {"side": "source"},
                {"side": "target"},
                object(),
                max_workers=1,
            )

        self.assertEqual("broken.md", result[0]["file"])
        self.assertFalse(result[0]["match"])
        self.assertIn("failed to fetch broken.md", result[0]["issue"])

    def test_collect_document_structures_fetches_once_and_splits_results(self):
        contents = {
            ("source", "wrapped.md"): '<CustomContent plan="essential">\n</CustomContent>\n',
            ("target", "wrapped.md"): '<CustomContent plan="essential">\n</CustomContent>\n',
            ("source", "heading.md"): "# Title\n\n## Step\n",
            ("target", "heading.md"): "# 标题\n\n### 步骤\n",
        }

        def fake_fetch(ctx, filepath, github_client):
            return contents[(ctx["side"], filepath)]

        with mock.patch.object(verify_translation, "_fetch_content", side_effect=fake_fetch) as fetch:
            custom_data, heading_data = verify_translation.collect_document_structures(
                ["wrapped.md", "heading.md"],
                ["heading.md"],
                {"side": "source"},
                {"side": "target"},
                object(),
                max_workers=2,
            )

        self.assertEqual(4, fetch.call_count)
        self.assertEqual(["wrapped.md", "heading.md"], [item["file"] for item in custom_data])
        self.assertEqual(["heading.md"], [item["file"] for item in heading_data])
        self.assertTrue(custom_data[0]["match"])
        self.assertFalse(heading_data[0]["match"])

    def test_custom_content_sheet_omits_files_with_no_custom_content(self):
        plain = _custom_content_data("plain.md", "# Plain\n", "# 普通\n")
        wrapped = _custom_content_data(
            "wrapped.md",
            '<CustomContent plan="essential">\n\n# Guide\n\n</CustomContent>\n',
            '<CustomContent plan="essential">\n\n# 指南\n\n</CustomContent>\n',
        )

        with tempfile.TemporaryDirectory() as tmpdir:
            output = os.path.join(tmpdir, "report.xlsx")
            write_excel(
                [_row("plain.md"), _row("wrapped.md")],
                output,
                "source",
                "target",
                5,
                custom_content_data=[plain, wrapped],
            )

            workbook = load_workbook(output)
            self.assertIn("Structure Details", workbook.sheetnames)
            sheet = workbook["CustomContent"]
            values = [
                cell
                for row in sheet.iter_rows(values_only=True)
                for cell in row
                if isinstance(cell, str)
            ]

        self.assertNotIn("plain.md", values)
        self.assertIn("wrapped.md", values)

    def test_custom_content_sheet_is_not_created_when_all_files_have_no_custom_content(self):
        plain = _custom_content_data("plain.md", "# Plain\n", "# 普通\n")

        with tempfile.TemporaryDirectory() as tmpdir:
            output = os.path.join(tmpdir, "report.xlsx")
            write_excel(
                [_row("plain.md")],
                output,
                "source",
                "target",
                5,
                custom_content_data=[plain],
            )

            workbook = load_workbook(output)

        self.assertNotIn("CustomContent", workbook.sheetnames)


if __name__ == "__main__":
    unittest.main()
